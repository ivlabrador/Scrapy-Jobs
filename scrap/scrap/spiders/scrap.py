import scrapy
from openpyxl import Workbook

lista = []

class JobSpider(scrapy.Spider):
    name = 'jobspider'
    start_urls = [
        'https://jobtoday.com/es/trabajos-python/barcelona?page=1',
        'https://jobtoday.com/es/trabajos-python/barcelona?page=2',
        'https://jobtoday.com/es/trabajos-python/barcelona?page=3',
        'https://jobtoday.com/es/trabajos-python/barcelona?page=4',
        'https://jobtoday.com/es/trabajos-python/barcelona?page=5',
        'https://jobtoday.com/es/trabajos-python/barcelona?page=6',
        'https://jobtoday.com/es/trabajos-python/barcelona?page=7',
        'https://jobtoday.com/es/trabajos-python/barcelona?page=8',
        'https://jobtoday.com/es/trabajos-python/barcelona?page=9',
        'https://jobtoday.com/es/trabajos-python/barcelona?page=10',
    ]
    lista = []

    def make_workbook(self, lista):
        wb = Workbook()
        ws = wb.create_sheet(title='Barcelona Python Jobs')
        headers = ['LINK', 'TITULO', 'EMPRESA', 'PUBLICADO', 'LUGAR', 'DESCRIPCION']
        columns = len(lista[0])
        rows = len(lista)
        for row in range(1, rows+1):
            for col in range(1, columns+1):
                if row == 1:
                    _ = ws.cell(column=col, row=row, value=headers[col-1])
                else:
                    _ = ws.cell(column=col, row=row, value=lista[row-2][col-1])

        wb.save(filename='data.xlsx')

    def parse(self, response):
        elements = response.css("li.flex")
        for e in elements:
            if e.css("strong::text").get() != None:
                element = []
                link = e.css("a::attr(href)").get()
                element.append(link)
                title = e.css("strong::text").get()
                element.append(title)
                company = e.css("span.not-italic::text").get()
                element.append(company)
                public = e.css("span.text-jt-gray-400::text").get()
                element.append(public)
                place = e.css("span.text-jt-gray-400.line-clamp-1::text").get()
                element.append(place)
                description = e.css("p::text").get()
                element.append(description)
                lista.append(element)
            else:
                pass

        excel = self.make_workbook(lista)
        return excel
